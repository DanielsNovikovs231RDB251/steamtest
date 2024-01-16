import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from openpyxl import Workbook, load_workbook 
service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)

wb=load_workbook('Data.xlsx')
ws=wb['Category']
Max_Row=ws.max_row
Cat_List=[]
for i in range(2, Max_Row):
    Cat_List.append((ws['b' + str(i)].value))
print("Choose game category:")
print("(Enter 'List' to look at catagories)")
while True:
    Category = input()
    if Category == "List":
        print(' | '.join(map(str, Cat_List)))
    elif Category in Cat_List:
        break
    else:
        print("There is no such category")
        print("(Enter 'List' to look at catagories)")

url = "https://store.steampowered.com"
driver.get(url)
time.sleep(2)
find_input = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.ID, "input"))
    )
find_input.send_keys(Category)

    #find_output = WebDriverWait(driver, 10).until(
     #   EC.presence_of_element_located((By.ID, "output"))
   # )
   # code.append(find_output.get_attribute("value"))